import openpyxl
import openpyxl.utils

wb_matome = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ.xlsx")

wb_jinji = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_人事部.xlsx")

month_list = ["4月","5月","6月"]

for month in month_list:
    ws_matome = wb_matome[month]
    ws_jinji = wb_jinji[month]


    for i in range(ws_jinji.max_column-1):
        ws_matome.cell(column= 3+i,row= 2).value = ws_jinji.cell(row=2,column=2+i).value
        ws_matome.cell(column= 3+i,row= 3).value = ws_jinji.cell(row=3,column=2+i).value


# # 4月分の処理
# for i in range(30):
#     ws_matome4.cell(column= 3+i,row= 2).value = ws_jinji4.cell(row=2,column=2+i).value
#     ws_matome4.cell(column= 3+i,row= 3).value = ws_jinji4.cell(row=3,column=2+i).value
    

# # 5月分の処理
# for i in range(30):
#     ws_matome5.cell(column= 3+i,row= 2).value = ws_jinji5.cell(row=2,column=2+i).value
#     ws_matome5.cell(column= 3+i,row= 3).value = ws_jinji5.cell(row=3,column=2+i).value
    

# # 6月分の処理

# for i in range(30):
#     ws_matome6.cell(column= 3+i,row= 2).value = ws_jinji6.cell(row=2,column=2+i).value
#     ws_matome6.cell(column= 3+i,row= 3).value = ws_jinji6.cell(row=3,column=2+i).value
    


wb_matome.save("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ２.xlsx")

openpyxl.utils.column_index_from_string("AE")
openpyxl.utils.get_column_letter(31)


