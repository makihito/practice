import openpyxl


import openpyxl

wb = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_人事部.xlsx")

print (wb)

ws = wb["4月"]

print(ws)

ws.cell(row=2,column=2).value

ws.cell(row=5,column=5).value = 1000

# 名前をつけて保存
wb.save("test.xlsx")

# 上書き保存
wb.save("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_人事部.xlsx")
