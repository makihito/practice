
import openpyxl

wb_jinji = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_人事部.xlsx")

wb_matome = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ.xlsx")

ws_jinji = wb_jinji["4月"]
ws_matome = wb_matome["4月"]

for i in range(30):
    ws_matome.cell(row=2,column=3+i).value = ws_jinji.cell(row=2,column=2+i).value

x = 2 + 3

wb_matome.save("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ2.xlsx")

