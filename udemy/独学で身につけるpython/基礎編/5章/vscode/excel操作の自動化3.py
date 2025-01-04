import openpyxl
import openpyxl.utils

wb_matome = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ.xlsx")


month_list = ["4月","5月","6月"]
kakubu_list = ["人事部","企画部","営業部"]

for kakubu in kakubu_list:
    wb_kakubu = openpyxl.load_workbook("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_{}.xlsx".format(kakubu))

    # if kakubu == "人事部":
    #     shussha_row = 2
    #     zaitaku_row = 3
    # if kakubu == "企画部":
    #     shussha_row = 4
    #     zaitaku_row = 5
    # if kakubu == "営業部":
    #     shussha_row = 6
    #     zaitaku_row = 7



    for month in month_list:
        ws_matome = wb_matome[month]
        ws_kakubu = wb_kakubu[month]
        
        for j in range(1,ws_matome.max_row+1):
            if ws_matome.cell(column= 1,row=j).value == kakubu:
                shussha_row = j
                zaitaku_row = j+1


        for i in range(ws_kakubu.max_column-1):
            ws_matome.cell(column= 3+i,row=shussha_row ).value = ws_kakubu.cell(row=2,column=2+i).value
            ws_matome.cell(column= 3+i,row=zaitaku_row ).value = ws_kakubu.cell(row=3,column=2+i).value


wb_matome.save("/Users/maetaakihito/programing/github/udemy/独学で身につけるpython基礎編/5章/出社在宅集計表_まとめ２.xlsx")